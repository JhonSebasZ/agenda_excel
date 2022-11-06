from openpyxl import *
from openpyxl import Workbook
from datetime import date

from Contacto import Contacto

class Agenda:
    listaContactos = []

    def guardar(self, contacto):
        """
        Guardar Contacto

        :param contacto: Datos del nuevo contacto
        """       
        self.escribirContacto(contacto)  # Escribe el contactos en un archivo

    def modificar(self, index, contactoModificado):
        """
        Modificar Contacto

        :param index: identificador del contacto que se quire modificar 
        :param contactoModificado: Contacto con los datos modificados
        """
        book = load_workbook('bd_contactos.xlsx')
                
        sheet = book.active
        sheet[f'B{index+1}'] = contactoModificado[0]
        
        sheet = book.active 
        sheet[f'C{index+1}'] = contactoModificado[1]
        
        sheet = book.active 
        sheet[f'D{index+1}'] = contactoModificado[2]
        
        sheet = book.active 
        sheet[f'E{index+1}'] = contactoModificado[3]
        
        sheet = book.active 
        sheet[f'F{index+1}'] = date.today()
        
        book.save('bd_contactos.xlsx')

    def eliminar(self, index):
        """
        Eliminar contacto

        :param index: identificador del contacto que se quiere eliminar
        """
        book = load_workbook('bd_contactos.xlsx')
        sheet = book.active
        
        sheet.delete_rows(index+1)
        
        max_row = book.active.max_row
        for i in range(index+1,max_row+1):
            sheet = book.active
            sheet[f'A{i}'] = sheet[f'A{i}'].value - 1
            
        book.save('bd_contactos.xlsx')

    def buscar(self, caracteres):
        """
        Buscar Contacto

        Busca contactos que coincida con los caracteres ingresados

        :param caracteres: caracteres que se comparan con el nombre de cada contacto

        :return: Lista de contactos encontrados
        """
        book = load_workbook('bd_contactos.xlsx')
        max_row = book.active.max_row
        sheet = book.active
        contactosEncontrados = []
        for i in range(2, max_row+1):
            nombreconfi = sheet[f'B{i}']
            if caracteres.capitalize() in nombreconfi.value:
                contacto = []
                contacto.append(sheet[f'A{i}'].value)
                contacto.append(sheet[f'B{i}'].value)
                contacto.append(sheet[f'C{i}'].value)
                contacto.append(sheet[f'D{i}'].value)
                contacto.append(sheet[f'E{i}'].value)
                contactosEncontrados.append(contacto)
        return contactosEncontrados

    def cargar(self):
        """
        Cargar contactos

        Carga todos los contactos que tenga el archivo bd_contactos.xlsx y los almacena en la lista listaContactos
        
        :return: list
        """
        listaContactos = []
        book = load_workbook('bd_contactos.xlsx')
        max_row = book.active.max_row
        
        for i in range(2, max_row+1):
            contacto = []
            sheet = book.active
            contacto.append(sheet[f'A{i}'].value)
            contacto.append(sheet[f'B{i}'].value)
            contacto.append(sheet[f'C{i}'].value)
            contacto.append(sheet[f'D{i}'].value)
            contacto.append(sheet[f'E{i}'].value)
            listaContactos.append(contacto)
        return listaContactos
    
    def escribirContacto(self, contacto):
        """
        Escribir Contacto

        :param contacto: Datos del nuevo contacto
        
        Añade un nuevo contacto en la bd_contactos.xlsx
        """
        book = load_workbook('bd_contactos.xlsx')
        max_row = book.active.max_row
        
        sheet = book.active
        sheet[f'A{max_row+1}'] = max_row
        
        sheet = book.active
        sheet[f'B{max_row+1}'] = contacto.nombre
        
        sheet = book.active 
        sheet[f'C{max_row+1}'] = contacto.apellido
        
        sheet = book.active 
        sheet[f'D{max_row+1}'] = contacto.telefono
        
        sheet = book.active 
        sheet[f'E{max_row+1}'] = contacto.correo
        
        sheet = book.active 
        sheet[f'F{max_row+1}'] = date.today()
        
        book.save('bd_contactos.xlsx')
        
    
